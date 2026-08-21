"""Turning uploaded files into something the model can read.

One internal format: PDF. Documents and presentations go through headless
LibreOffice; spreadsheets become Markdown tables because rendering a table as a
picture of a table reads worse and costs more; images and text pass through.

The concurrency rule here is not an optimisation, it is a correctness fix.
Measured: six concurrent conversions sharing one user profile produced two PDFs
and four bare exit-1s with **no error output at all**. LibreOffice does not
serialise on a shared profile — it fails, and says nothing.
"""

from __future__ import annotations

import asyncio
import os
import subprocess
import tempfile
from pathlib import Path

SOFFICE = os.environ.get("AI_ANKI_SOFFICE", "soffice")

CONVERTIBLE = {".docx", ".doc", ".pptx", ".ppt", ".odt", ".odp", ".rtf"}
SPREADSHEET = {".xlsx", ".xlsm", ".xls", ".csv"}
IMAGE = {".png", ".jpg", ".jpeg", ".gif", ".webp"}

# Sized to processors, and deliberately unrelated to the Claude fan-out limit.
# On a single vCPU, concurrent conversion buys nothing and multiplies memory:
# measured 2922ms for one, 12280ms for four. At ~218MB peak each, that is also
# how a 1GB machine runs out of memory.
MAX_CONCURRENT = max(1, min(4, (os.cpu_count() or 1)))
_semaphore = asyncio.Semaphore(MAX_CONCURRENT)

# Generous, because it bounds a hung process rather than a slow one. A 40-slide
# deck converts in about 3 seconds.
TIMEOUT_SECONDS = 180


class ConversionFailed(Exception):
    """A source could not be converted. Names the file, because the user has to act."""


def needs_conversion(path: Path) -> bool:
    return path.suffix.lower() in CONVERTIBLE


def is_spreadsheet(path: Path) -> bool:
    return path.suffix.lower() in SPREADSHEET


def is_image(path: Path) -> bool:
    return path.suffix.lower() in IMAGE


def convert_to_pdf(source: Path, destination_dir: Path) -> Path:
    """Convert one file to PDF. Blocking; call it off the event loop.

    Each invocation gets its own throwaway user-installation profile. This is
    mandatory rather than tidy: sharing one is the silent-failure case above,
    and a missing one fails fast with exit 77 and a clear message.
    """
    destination_dir.mkdir(parents=True, exist_ok=True)
    with tempfile.TemporaryDirectory(prefix="soffice-profile-") as profile:
        completed = subprocess.run(
            [
                SOFFICE,
                f"-env:UserInstallation=file://{profile}",
                "--headless",
                "--norestore",
                "--convert-to",
                "pdf",
                "--outdir",
                str(destination_dir),
                str(source),
            ],
            capture_output=True,
            text=True,
            timeout=TIMEOUT_SECONDS,
        )

    produced = destination_dir / f"{source.stem}.pdf"
    # The output file is the only trustworthy signal. LibreOffice can exit 0
    # having produced nothing, and can fail having said nothing, so neither the
    # exit code nor stderr is checked first.
    if not produced.exists():
        raise ConversionFailed(
            f"Could not convert {source.name}. LibreOffice exited "
            f"{completed.returncode} without producing a PDF. "
            f"{(completed.stderr or '').strip()[:200]}"
        )
    return produced


async def convert_many(sources: list[Path], destination_dir: Path) -> list[Path]:
    """Convert a batch, bounded by the semaphore."""

    async def one(source: Path) -> Path:
        async with _semaphore:
            return await asyncio.to_thread(convert_to_pdf, source, destination_dir)

    return list(await asyncio.gather(*(one(source) for source in sources)))


def spreadsheet_to_markdown(path: Path) -> str:
    """Read a spreadsheet as tables.

    Rendered to pages it would be billed as text *and* image, and read worse.
    As Markdown it is just text, and the structure survives.
    """
    from openpyxl import load_workbook

    workbook = load_workbook(path, read_only=True, data_only=True)
    out: list[str] = []
    for sheet in workbook.worksheets:
        rows = [
            ["" if cell is None else str(cell) for cell in row]
            for row in sheet.iter_rows(values_only=True)
        ]
        rows = [row for row in rows if any(cell.strip() for cell in row)]
        if not rows:
            continue
        out.append(f"## {sheet.title}\n")
        header, *body = rows
        out.append("| " + " | ".join(header) + " |")
        out.append("| " + " | ".join("---" for _ in header) + " |")
        for row in body:
            padded = row + [""] * (len(header) - len(row))
            out.append("| " + " | ".join(padded[: len(header)]) + " |")
        out.append("")
    workbook.close()
    return "\n".join(out)
